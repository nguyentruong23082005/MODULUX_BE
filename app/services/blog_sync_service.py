from __future__ import annotations

import logging
import threading
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from io import BytesIO

from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import BLOG_SYNC_INTERVAL_HOURS, BLOG_SYNC_SCHEDULER_ENABLED
from app.db.session import SessionLocal, engine
from app.models.blog_sync import Blog, SyncLog
from app.repositories.blog_sync_repository import BlogSyncRepository
from app.services.blog_crawler import CrawledBlogItem, CrawledBlogLink, ExternalBlogCrawler

logger = logging.getLogger(__name__)

BLOG_SYNC_LOCK = threading.Lock()
BLOG_SYNC_SCHEDULER: BackgroundScheduler | None = None


@dataclass(slots=True)
class SyncSummary:
    total: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0

    def to_dict(self) -> dict[str, int]:
        return asdict(self)


class BlogSyncService:
    def __init__(
        self,
        db: Session,
        *,
        repository: BlogSyncRepository | None = None,
        crawler: ExternalBlogCrawler | None = None,
    ) -> None:
        self.db = db
        self.repository = repository or BlogSyncRepository()
        self.crawler = crawler or ExternalBlogCrawler()

    def sync_all_blogs(self) -> SyncSummary:
        return self._run_locked_sync(self.crawler.crawl_blog_links())

    def import_from_excel(self, file_bytes: bytes) -> SyncSummary:
        urls = self._read_excel_urls(file_bytes)
        return self._run_locked_sync(urls)

    def _run_locked_sync(self, entries: list[CrawledBlogLink | str]) -> SyncSummary:
        acquired = BLOG_SYNC_LOCK.acquire(blocking=False)
        if not acquired:
            raise RuntimeError("Blog sync is already running")
        try:
            return self._sync_entries(entries)
        finally:
            BLOG_SYNC_LOCK.release()

    def _sync_entries(self, entries: list[CrawledBlogLink | str]) -> SyncSummary:
        unique_entries: list[CrawledBlogLink | str] = []
        seen_urls: set[str] = set()

        for entry in entries:
            url = entry.source_url if isinstance(entry, CrawledBlogLink) else entry
            if url in seen_urls:
                continue
            seen_urls.add(url)
            unique_entries.append(entry)

        summary = SyncSummary(total=len(unique_entries))

        for entry in unique_entries:
            url = entry.source_url if isinstance(entry, CrawledBlogLink) else entry
            try:
                item = self.crawler.crawl_blog_detail(
                    url,
                    blog_type=entry.blog_type if isinstance(entry, CrawledBlogLink) else None,
                )
                status = self.sync_blog_item(item)
                if status == "inserted":
                    summary.inserted += 1
                elif status == "updated":
                    summary.updated += 1
                else:
                    summary.skipped += 1
            except Exception:
                logger.exception("Failed to sync blog '%s'", url)
                self.db.rollback()

        sync_log = SyncLog(
            type="blog",
            total=summary.total,
            inserted=summary.inserted,
            updated=summary.updated,
            skipped=summary.skipped,
            created_at=datetime.now(timezone.utc),
        )
        self.repository.add_sync_log(self.db, sync_log)
        self.db.commit()
        return summary

    def sync_blog_item(self, item: CrawledBlogItem) -> str:
        now = datetime.now(timezone.utc)
        existing = self.repository.get_blog_by_source_url(self.db, item.source_url)

        if existing is None:
            blog = Blog(
                title=item.title,
                blog_type=item.blog_type,
                source_type=item.blog_type,
                type_override=None,
                slug=self._build_unique_slug(item.slug),
                content=item.content,
                source_url=item.source_url,
                source_hash=item.source_hash,
                created_at=item.source_created_at,
                updated_at=item.source_updated_at,
                last_synced_at=now,
            )
            self.repository.add_blog(self.db, blog)
            self.db.commit()
            return "inserted"

        if existing.source_hash == item.source_hash:
            existing.last_synced_at = now
            self.db.commit()
            return "skipped"

        existing.title = item.title
        existing.source_type = item.blog_type
        if existing.type_override:
            existing.blog_type = existing.type_override
        else:
            existing.blog_type = item.blog_type
        existing.slug = self._build_unique_slug(item.slug, exclude_blog_id=existing.id)
        existing.content = item.content
        existing.source_hash = item.source_hash
        existing.updated_at = item.source_updated_at
        existing.last_synced_at = now
        self.db.commit()
        return "updated"

    def _build_unique_slug(self, raw_slug: str, *, exclude_blog_id: int | None = None) -> str:
        base_slug = raw_slug.strip("-") or "blog"
        slug = base_slug
        counter = 2
        while self.repository.slug_exists(self.db, slug, exclude_blog_id=exclude_blog_id):
            slug = f"{base_slug}-{counter}"
            counter += 1
        return slug

    def _read_excel_urls(self, file_bytes: bytes) -> list[str]:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")

        header_row = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        if "url" not in header_row:
            raise ValueError("Excel file must contain a 'url' column")

        url_index = header_row.index("url")
        urls: list[str] = []
        for row in rows[1:]:
            if row is None or url_index >= len(row):
                continue
            value = row[url_index]
            if value is None:
                continue
            url = str(value).strip()
            if url:
                urls.append(url)

        if not urls:
            raise ValueError("Excel file does not contain any blog URLs")

        return urls


def is_blog_sync_running() -> bool:
    return BLOG_SYNC_LOCK.locked()


def ensure_blog_sync_schema() -> None:
    statements = [
        "ALTER TABLE blogs ADD COLUMN IF NOT EXISTS type VARCHAR(50) NOT NULL DEFAULT 'PROJECTS'",
        "ALTER TABLE blogs ADD COLUMN IF NOT EXISTS source_type VARCHAR(50) NOT NULL DEFAULT 'PROJECTS'",
        "ALTER TABLE blogs ADD COLUMN IF NOT EXISTS type_override VARCHAR(50)",
        "UPDATE blogs SET source_type = type WHERE source_type IS NULL OR source_type = ''",
        "CREATE INDEX IF NOT EXISTS ix_blogs_type ON blogs (type)",
        "CREATE INDEX IF NOT EXISTS ix_blogs_source_type ON blogs (source_type)",
    ]
    with engine.begin() as connection:
        for statement in statements:
            connection.execute(text(statement))


def run_blog_sync_job() -> None:
    db = SessionLocal()
    try:
        BlogSyncService(db).sync_all_blogs()
    except RuntimeError:
        logger.info("Skipped background sync because another sync is already running")
    except Exception:
        logger.exception("Unexpected error during background blog sync")
        db.rollback()
    finally:
        db.close()


def start_blog_sync_scheduler() -> None:
    global BLOG_SYNC_SCHEDULER

    if not BLOG_SYNC_SCHEDULER_ENABLED or BLOG_SYNC_SCHEDULER is not None:
        return

    scheduler = BackgroundScheduler(timezone="UTC")
    scheduler.add_job(
        run_blog_sync_job,
        trigger="interval",
        hours=BLOG_SYNC_INTERVAL_HOURS,
        id="blog-sync-job",
        replace_existing=True,
        max_instances=1,
    )
    scheduler.start()
    BLOG_SYNC_SCHEDULER = scheduler
    logger.info("Blog sync scheduler started with %s-hour interval", BLOG_SYNC_INTERVAL_HOURS)


def stop_blog_sync_scheduler() -> None:
    global BLOG_SYNC_SCHEDULER

    if BLOG_SYNC_SCHEDULER is None:
        return

    BLOG_SYNC_SCHEDULER.shutdown(wait=False)
    BLOG_SYNC_SCHEDULER = None
    logger.info("Blog sync scheduler stopped")
